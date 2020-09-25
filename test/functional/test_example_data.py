from time import strftime
from typing import Union

import pytest
import eccpy
from tempfile import TemporaryDirectory
from eccpy.tools import get_eccpy_module_path
from openpyxl import load_workbook
from pathlib import Path
from shutil import copyfile


@pytest.mark.regression
def test_processing_of_example_data():
    eccpy_module_path: Union[Path, str] = get_eccpy_module_path()

    with TemporaryDirectory() as tmpdirname:
        # uncomment to view test output in repo dir during development
        #tmpdirname: Union[Path, str] = eccpy_module_path / "test"

        tmpdir: Path = Path(tmpdirname)
        temp_path_for_testing: Union[Path, str] = tmpdir / "eccpy_temp_test_output"
        print(f" Test output will be saved to {temp_path_for_testing}, and deleted after tests are finished")

        test_settings: Union[Path, str] = set_up_excel_settings_for_functional_test(temp_path_for_testing)

        eccpy.run_curvefit(test_settings)
        eccpy.run_gatherer(test_settings)

        assert_curvefit_output_files_exist(temp_path_for_testing)

        assert_gatherer_output_files_exist(temp_path_for_testing)

    # assert temporary output files are removed
    assert not Path(tmpdirname).is_dir()


def assert_gatherer_output_files_exist(temp_path_for_testing):
    date_string = strftime("%Y%m%d")
    assert (temp_path_for_testing / f"analysed/{date_string}").is_dir()
    assert (temp_path_for_testing / f"analysed/{date_string}/{date_string}.xlsx").is_file()
    assert (temp_path_for_testing / f"analysed/{date_string}/{date_string}_datapoints.png").is_file()


def assert_curvefit_output_files_exist(temp_path_for_testing):
    assert (temp_path_for_testing / "analysed").is_dir()
    assert (temp_path_for_testing / f"generated_data_0/curves/AA control.png").is_file()
    assert (temp_path_for_testing / f"generated_data_0/generated_data_0_summary.png").is_file()


def set_up_excel_settings_for_functional_test(test_temp_output_path):
    eccpy_module_path = get_eccpy_module_path()

    orig_settings_xlsx: Union[Path, str] = eccpy_module_path / "eccpy/examples/example_settings/ECCpy_settings_template.xlsx"
    temp_test_settings_xlsx: Union[Path, str] = test_temp_output_path / " settings_example_data.xlsx"

    example_data_dir: Path = eccpy_module_path / "eccpy/examples/example_data"
    #output_data_dir: str = str(eccpy_module_path / "test/temp_output")

    if not temp_test_settings_xlsx.parent.is_dir():
        temp_test_settings_xlsx.parent.mkdir(parents=True)

    copyfile(orig_settings_xlsx, temp_test_settings_xlsx)
    assert temp_test_settings_xlsx.is_file()

    wb = load_workbook(temp_test_settings_xlsx)
    sheet = "files"
    input_data_cells = ["G2", "G3", "G4", "G5"]
    output_data_cells = ["H2", "H3", "H4", "H5"]

    ws = wb[sheet]

    for cell in input_data_cells:
        ws[cell] = str(example_data_dir)
    for cell in output_data_cells:
        ws[cell] = str(test_temp_output_path)

    wb.save(temp_test_settings_xlsx)

    return temp_test_settings_xlsx


