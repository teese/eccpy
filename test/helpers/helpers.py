from pathlib import Path
from shutil import copyfile
from typing import Union

from openpyxl import load_workbook

from eccpy.tools import get_eccpy_module_path


def set_up_excel_settings_for_functional_test(temp_path_for_testing: Path, replacement_dict: dict) -> Path:
    eccpy_module_path = get_eccpy_module_path()

    orig_settings_xlsx: Union[Path, str] = eccpy_module_path / "eccpy/examples/example_settings/ECCpy_settings_template.xlsx"
    temp_test_settings_xlsx: Union[Path, str] = temp_path_for_testing / " settings_example_data.xlsx"

    if not temp_test_settings_xlsx.parent.is_dir():
        temp_test_settings_xlsx.parent.mkdir(parents=True)

    copyfile(orig_settings_xlsx, temp_test_settings_xlsx)
    assert temp_test_settings_xlsx.is_file()

    replace_xlsx_cells_by_replacement_dict(replacement_dict, temp_test_settings_xlsx)

    return temp_test_settings_xlsx


def replace_xlsx_cells_by_replacement_dict(replacement_dict: dict, xlsx_path: Union[Path, str]):
    wb = load_workbook(xlsx_path)
    for sheet, replacements in replacement_dict.items():
        ws = wb[sheet]
        for cell, value in replacements.items():
            ws[cell] = value
    wb.save(xlsx_path)


def get_replacement_dict_with_test_paths(temp_path_for_testing, example_data_dir=None):
    eccpy_module_path = get_eccpy_module_path()
    if not example_data_dir:
        example_data_dir: Path = eccpy_module_path / "eccpy/examples/example_data"
    replacement_dict = {}
    files_replacements: dict = {}
    input_data_cells = ["G2", "G3", "G4", "G5"]
    output_data_cells = ["H2", "H3", "H4", "H5"]
    for cell in input_data_cells:
        files_replacements[cell] = str(example_data_dir)
    for cell in output_data_cells:
        files_replacements[cell] = str(temp_path_for_testing)
    replacement_dict["files"] = files_replacements

    return replacement_dict
